import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
from ..models.seating_plan import SeatingPlan
from .alphanum_handler import alphanum_sort_key

def import_excel_to_plan(file_path: str, plan: SeatingPlan) -> None:
    """Parses an Excel file and populates the provided SeatingPlan object."""
    wb = load_workbook(filename=file_path, read_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_indices = {header: idx for idx, header in enumerate(headers) if header is not None}

    required_headers = {"section", "rows", "seats", "type"}
    if not required_headers.issubset(header_indices.keys()):
        raise ValueError(f"Excel file must contain headers: {', '.join(required_headers)}")

    for row in ws.iter_rows(min_row=2):
        section_name = row[header_indices["section"]].value
        row_identifier = row[header_indices["rows"]].value
        seats_str = row[header_indices["seats"]].value
        type_value = row[header_indices["type"]].value
        
        if section_name is None or type_value is None:
            continue
        
        is_ga = int(type_value) != 0
        if section_name not in plan.sections:
            plan.add_section(section_name, is_ga=is_ga)

        if not is_ga and seats_str:
            seat_labels = [s.strip() for s in str(seats_str).split(",") if s.strip()]
            for seat_label in seat_labels:
                plan.sections[section_name].add_seat(str(row_identifier), seat_label)

def import_avail_xml_to_plan(file_path: str, plan: SeatingPlan) -> None:
    """Parses Avail XML content and populates the provided SeatingPlan object."""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    soup = BeautifulSoup(content, 'lxml')
    section_id_list = soup.find('section_id_list')
    if not section_id_list:
        return

    for e in section_id_list.find_all('e'):
        if e.find('section_id') is None:
            continue
            
        sec_name = e.find('section_name').text.strip()
        is_ga = e.find('is_ga').text.lower() == 'true'
        
        if sec_name not in plan.sections:
            plan.add_section(sec_name, is_ga=is_ga)
            
        if not is_ga:
            row_names = [r.text.strip() for r in e.find('row_names').find_all('e')]
            seat_names = [s.text.strip() for s in e.find('seat_names').find_all('e')]
            
            for r in row_names:
                for s in seat_names:
                    plan.sections[sec_name].add_seat(r, s)

def export_plan_to_excel(file_path: str, plan: SeatingPlan) -> None:
    """Exports the SeatingPlan to an Excel file compatible with the importer."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Seating Plan"
    ws.append(["section", "rows", "seats", "secnam", "capacity", "type"])

    for section in plan.sections.values():
        if section.is_ga:
            ws.append([section.name, "", "", section.name, "1", 1])
            continue
            
        rows = {}
        for seat in section.seats.values():
            rows.setdefault(seat.row_number, []).append(str(seat.seat_number))

        for row_number, seat_list in rows.items():
            # Sort seats using the robust alphanumeric sort key
            seat_list_sorted = sorted(seat_list, key=alphanum_sort_key)

            ws.append([
                section.name,
                row_number,
                ",".join(seat_list_sorted),
                section.name,
                "",
                0
            ])
    wb.save(file_path)