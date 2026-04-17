"""Excel format importer for seating plans."""

from pathlib import Path
from typing import Optional, TYPE_CHECKING

if TYPE_CHECKING:
    from ...domain.models.seating_plan import SeatingPlan

from openpyxl import load_workbook

from ...domain.exceptions import SeatingPlanException
from .abstract import Importer


class ExcelImporter(Importer):
    """Import seating plans from Excel (.xlsx) files."""
    
    SUPPORTED_EXTENSIONS = (".xlsx",)
    REQUIRED_HEADERS = {"section", "rows", "seats", "type"}
    
    def supports(self, file_path: str) -> bool:
        """Check if this importer handles the file type."""
        return file_path.lower().endswith(self.SUPPORTED_EXTENSIONS)
    
    def import_seating_plan(self, file_path: str, plan_name: str) -> "SeatingPlan":
        """Import a seating plan from an Excel file.
        
        Excel file format:
        - Headers: section, rows, seats, type (required)
        - type: 0 = numbered seats, non-zero = general admission
        - seats: comma-separated seat labels for numbered sections
        
        Args:
            file_path: Path to the Excel file
            plan_name: Name for the imported seating plan
            
        Returns:
            A populated SeatingPlan instance
            
        Raises:
            SeatingPlanException: If import fails
        """
        try:
            from ...domain.models.seating_plan import SeatingPlan
            
            file_path_obj = Path(file_path)
            
            if not file_path_obj.exists():
                raise SeatingPlanException(
                    f"File not found: {file_path}",
                    error_code="FILE_NOT_FOUND"
                )
            
            # Load workbook
            wb = load_workbook(filename=str(file_path_obj), read_only=True)
            ws = wb.active
            print(DEBUG := f"Loaded workbook: {file_path}, active sheet: {ws.title if ws else 'None'}")
            if not ws:
                raise SeatingPlanException(
                    "No active worksheet in Excel file",
                    error_code="NO_WORKSHEET"
                )
            
            # Extract headers
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            header_indices = {
                header: idx 
                for idx, header in enumerate(headers) 
                if header is not None
            }
            
            # Verify required headers
            missing = self.REQUIRED_HEADERS - set(header_indices.keys())
            if missing:
                raise SeatingPlanException(
                    f"Excel file missing required columns: {', '.join(missing)}",
                    error_code="MISSING_HEADERS"
                )
            
            # Create seating plan
            seating_plan = SeatingPlan(plan_name or "Imported Plan")
            
            # Parse data rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                try:
                    section_name = row[header_indices["section"]].value
                    row_identifier = row[header_indices["rows"]].value
                    splitted_row_identifier = str(row_identifier).split(",") if row_identifier else []
                    seats_str = row[header_indices["seats"]].value
                    type_value = row[header_indices["type"]].value
                    
                    # Skip empty rows
                    if section_name is None or type_value is None:
                        continue
                    
                    # Determine if general admission
                    is_ga = int(type_value) != 0
                    
                    # Create section if needed
                    if section_name not in seating_plan.sections:
                        seating_plan.add_section(section_name, is_ga=is_ga)
                    
                    # Add seats for numbered sections
                    if not is_ga and seats_str:
                        seat_labels = [
                            s.strip() 
                            for s in str(seats_str).split(",") 
                            if s.strip()
                        ]
                        for row_id in splitted_row_identifier:
                            for seat_label in seat_labels:
                                seating_plan.sections[section_name].add_seat(
                                    str(row_id), 
                                    seat_label
                                )
                
                except (ValueError, IndexError) as e:
                    raise SeatingPlanException(
                        f"Invalid data in row {row_idx}: {e}",
                        error_code="INVALID_ROW_DATA"
                    ) from e
            
            wb.close()
            return seating_plan
            
        except SeatingPlanException:
            raise
        except Exception as e:
            raise SeatingPlanException(
                f"Failed to import Excel file: {e}",
                error_code="IMPORT_FAILED"
            ) from e