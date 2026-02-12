import json
import re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from typing import Dict, List
from .section import Section

class MergeConflictError(Exception):
    """Raised when attempting to merge sections that contain conflicting seats."""

class SeatingPlan:
    """Represents an entire seating plan project with multiple sections."""

    def __init__(self, name: str = "Unnamed Plan") -> None:
        self.sections: Dict[str, Section] = {}
        self.name: str = name or "Unnamed Plan"

    # ---- Section Manipulation ----
    def add_section(self, name: str, is_ga: bool = False) -> None:
        if name not in self.sections:
            self.sections[name] = Section(name, is_ga=is_ga)
            self.sections[name].is_ga = is_ga

    def delete_section(self, name: str) -> None:
        self.sections.pop(name, None)

    def rename_section(self, old_name: str, new_name: str) -> None:
        if old_name in self.sections and new_name not in self.sections:
            section = self.sections.pop(old_name)
            section.rename(new_name)
            self.sections[new_name] = section

    def clone_section(self, name: str, new_name: str) -> None:
        if name in self.sections and new_name not in self.sections:
            cloned = self.sections[name].clone()
            cloned.name = new_name
            self.sections[new_name] = cloned

    def clone_section_many(self, name: str, count: int) -> List[str]:
        """
        Clone the given section 'count' times, returning a list of created section names.

        Naming strategy:
        - If the source name ends with a numeric suffix (e.g. "Section 1"), the clones will
          continue numbering from that suffix +1 (e.g. 2,3,...).
        - If the source name has no numeric suffix, clones will be created as "Name 2",
          "Name 3", ... (starting numbering at 2).
        - If a generated candidate name already exists, it is skipped by advancing the number
          until an unused name is found. Exactly 'count' new sections will be created.
        """
        created: List[str] = []
        if name not in self.sections or count <= 0:
            return created

        # Try to find a prefix and a start number. Example matches:
        #  - "I Ordine Palco 1" -> prefix="I Ordine Palco", start=1
        #  - "Balcony" -> prefix="Balcony", start=None -> we'll treat start as 1
        m = re.match(r"^(.*\S)(?:\s+(\d+))?$", name)
        if not m:
            prefix = name
            start_num = 1
        else:
            prefix = m.group(1) if m.group(1) else name
            start_num = int(m.group(2)) if m.group(2) else 1

        current = start_num + 1
        for _ in range(count):
            # find next unused candidate name
            candidate = f"{prefix} {current}"
            while candidate in self.sections:
                current += 1
                candidate = f"{prefix} {current}"
            # clone and register
            cloned = self.sections[name].clone()
            cloned.name = candidate
            self.sections[candidate] = cloned
            created.append(candidate)
            current += 1

        return created

    def merge_sections(self, source_section_names: List[str], new_section_name: str) -> None:
        """Create a new section containing all seats from the listed source sections.

        Conflicts (same row+seat present in more than one source) will abort the
        operation and raise MergeConflictError. The source sections are left
        unchanged.

        Args:
            source_section_names: list of existing section names to merge (>=2)
            new_section_name: name for the newly created section (must not exist)
        """
        # Validate input
        if not source_section_names or len(source_section_names) < 2:
            raise ValueError("Provide at least two source sections to merge")
        if new_section_name in self.sections:
            raise ValueError(f"Target section '{new_section_name}' already exists")

        # Collect source Section objects and validate existence
        sources: List[Section] = []
        for sname in source_section_names:
            if sname not in self.sections:
                raise KeyError(f"Source section '{sname}' does not exist")
            sources.append(self.sections[sname])

        # Build staging map and detect conflicts: seat_key -> list of source names
        seat_origins: Dict[str, List[str]] = {}
        for sec in sources:
            for key in sec.seats.keys():
                seat_origins.setdefault(key, []).append(sec.name)

        # Any seat key that appears more than once is a conflict
        conflicts = {k: v for k, v in seat_origins.items() if len(v) > 1}
        if conflicts:
            # Format a helpful message listing conflicting seat keys and origins
            msgs = [f"{k}: {','.join(v)}" for k, v in conflicts.items()]
            raise MergeConflictError("Conflicting seats detected: " + "; ".join(msgs))

        # No conflicts: create the new section and copy seats
        new_is_ga = all(sec.is_ga for sec in sources)
        new_section = Section(new_section_name, is_ga=new_is_ga)
        for sec in sources:
            for seat in sec.seats.values():
                new_section.add_seat(seat.row_number, seat.seat_number)

        # Register new section
        self.sections[new_section_name] = new_section

    # ---- Serialization ----
    def to_dict(self) -> dict:
        return {
            "seating_plan_name": self.name,
            "sections": [section.to_dict() for section in self.sections.values()]
        }

    def from_dict(self, data: dict) -> None:
        self.name = data.get("seating_plan_name", "Unnamed Plan")
        self.sections = {}
        for section_data in data.get("sections", []):
            section = Section.from_dict(section_data)
            self.sections[section.name] = section

    # ---- File I/O ----
    def export_project(self, file_path: str) -> None:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(self.to_dict(), f, indent=2, ensure_ascii=False)

    def import_project(self, file_path: str) -> None:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            self.from_dict(data)

    def import_from_excel(self, file_path: str) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(filename=file_path, read_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_indices = {header: idx for idx, header in enumerate(headers) if header is not None}

        required_headers = {"section", "rows", "seats", "type"}
        if not required_headers.issubset(header_indices.keys()):
            raise ValueError(f"Excel file must contain headers: {', '.join(required_headers)}")

        for row in ws.iter_rows(min_row=2):
            section_name = row[header_indices["section"]].value
            row_identifier = row[header_indices["rows"]].value
            seats_str = row[header_indices["seats"]].value
            type_value = row[header_indices["type"]].value
            if section_name is None or (row_identifier is None and (seats_str is None)) or type_value is None:
                continue
            
            if int(type_value) == 0:
                is_ga = False
            else:
                is_ga = True

            if section_name not in self.sections:
                self.add_section(section_name, is_ga=is_ga)

            try:
                seat_labels = [s.strip() for s in seats_str.split(",") if s.strip()]
            except Exception:
                continue
            for seat_label in seat_labels:
                self.sections[section_name].add_seat(row_identifier, seat_label)

    def import_from_avail(self, file_path: str) -> None:
       
        txt = file_path 
        with open(txt, 'r', encoding='utf-8') as f:
            content = f.read()
        
        def get_dict_per_row(tag):
            secname = tag.find('section_name')
            secname = secname.text
            seccode = tag.find( 'secnam_list')
            seccode = seccode.text
            row_tag = tag.find('row_names')
            row_nums = row_tag.find_all('e')
            row_nums = [row.text for row in row_nums]
            row_nums.sort()
            seats_tag = tag.find('seat_names')
            seat_nums = seats_tag.find_all('e')
            seat_nums = [seat.text for seat in seat_nums]
            seat_nums.sort()
            rows = ','.join(row_nums)
            seats = ','.join(seat_nums)
            isga = tag.find('is_ga')

            if isga.text == 'true':
                type = 1
                capacity = 1
                rows = None
                seats = None
            else:
                type = 0 
                capacity = None

            rows_dict = {'section' : secname, 'rows' : rows, 'seats' : seats, 'secnam' : seccode, 'type' : type, 'capacity' : capacity}
            rows_dict = {k: v.replace('\n', '').strip() if isinstance(v, str) else v for k, v in rows_dict.items()}

            return rows_dict

        def avail_parser( content: str) -> None:
            """Parse Avail XML content and populate the seating plan."""
            soup = BeautifulSoup(content,'lxml' )

            section_id_list = soup.find('section_id_list')

            section_tags = []

            for e in section_id_list.find_all('e'):
                if e.find('section_id') is not None:
                    section_tags.append(e)

            dicts = []
            for tag in section_tags:
                dict = get_dict_per_row(tag)

                dicts.append(dict)
            return dicts
        dicts = avail_parser(content)
        for row_dict in dicts:
            section_name = row_dict['section']
            row_identifier = row_dict['rows']
            seats_str = row_dict['seats']

            if section_name is None or row_identifier is None or seats_str is None:
                continue

            if section_name not in self.sections:
                self.add_section(section_name)

            row_labels = [r.strip() for r in row_identifier.split(",") if r.strip()]
            seat_labels = [s.strip() for s in seats_str.split(",") if s.strip()]
            for row_label in row_labels:
                for seat_label in seat_labels:
                    self.sections[section_name].add_seat(row_label, seat_label)

    def export_to_excel(self, file_path: str) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Seating Plan"

        headers = ["section", "rows", "seats", "secnam", "capacity", "type"]
        ws.append(headers)

        # Iterate through sections and rows
        for section in self.sections.values():
            if section.is_ga:
                ws.append([
                    section.name,           # section
                    "",                     # rows
                    "",                     # seats
                    section.name,           # secnam
                    "1",                    # capacity (set to 1)
                    1                       # type (1 for GA)
                ])
            rows = {}
            for seat in section.seats.values():
                rows.setdefault(seat.row_number, []).append(str(seat.seat_number))

            for row_number, seat_list in rows.items():
                try:
                    seat_list_sorted = sorted(seat_list, key=lambda x: int(x) if x.isdigit() else x)
                except:
                    seat_list_sorted = sorted(seat_list)
                ws.append([
                    section.name,             # section
                    row_number,               # rows
                    ",".join(seat_list_sorted),  # seats
                    section.name,             # secnam
                    "",                       # capacity (blank)
                    0                         # type  (0 for seated)
                ])
        wb.save(file_path)